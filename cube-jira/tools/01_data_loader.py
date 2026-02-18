"""
=================================================================
ЗАГРУЗЧИК ДАННЫХ В CUBE
=================================================================
Скрипт читает структуру данных, генерирует описания
на русском языке через GigaChat и создаёт YAML-модели для Cube.

Поддерживает режимы источника данных (database.driver в config.yml):
  - postgresql              — прямое подключение (psycopg2)
  - greenplum               — Greenplum через SQLAlchemy (Kerberos опционально)
  - hive                    — Hive через SQLAlchemy/PyHive (Kerberos)
  - duckdb                  — локальный DuckDB-файл
  - cube                    — чтение из работающего Cube API (без БД)

Knowledge Base (опционально):
  config.yml → knowledge_base_path: "./kb/jira_kb.yml"
  Внешний YAML-файл с описаниями таблиц, подсказками по колонкам
  и рекомендуемыми мерами для конкретного домена.

Дополнительно:
  --kb <file.yml>           — путь к Knowledge Base (переопределяет config.yml)
  --etl-plan <file.xlsx>    — обогатить модели из ETL execution plan файла
  --enrich-etl              — обогатить УЖЕ СУЩЕСТВУЮЩИЕ модели через ETL plan
  --enrich-with-llm         — при --enrich-etl переописать колонки через GigaChat
  --model-dir <dir>         — папка с моделями (для --enrich-etl)

Запуск:
  Полная генерация:
    python 01_data_loader.py
    python 01_data_loader.py --source cube
    python 01_data_loader.py --kb ./kb/jira_kb.yml --etl-plan plan.xlsx

  Обогащение существующих моделей (без перегенерации):
    python 01_data_loader.py --enrich-etl --etl-plan plan.xlsx
    python 01_data_loader.py --enrich-etl --etl-plan plan.xlsx --enrich-with-llm
    python 01_data_loader.py --enrich-etl --etl-plan plan.xlsx --model-dir ../model/cubes
=================================================================
"""

import os
import sys
import subprocess
import argparse
from pathlib import Path
from datetime import datetime

# ============================================================
# Автоустановка зависимостей
# ============================================================

def _ensure_packages():
    """Проверить и установить недостающие пакеты (базовые)"""
    required = {
        "yaml": "pyyaml",
        "langchain_gigachat": "langchain-gigachat",
    }
    missing = []
    for module, pip_name in required.items():
        try:
            __import__(module)
        except ImportError:
            missing.append(pip_name)
    if missing:
        print(f"📦 Установка недостающих пакетов: {', '.join(missing)}")
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "--quiet"] + missing
        )
        print("✅ Пакеты установлены")

_ensure_packages()

import yaml

# psycopg2 / duckdb подгружаются по необходимости в create_data_source()
try:
    import psycopg2
except ImportError:
    psycopg2 = None  # Не нужен для duckdb/cube режимов

# ============================================================
# Загрузка конфигурации
# ============================================================

def load_config(config_path="config.yml"):
    """Загрузить конфигурацию из YAML-файла"""
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


# ============================================================
# Подключение к GigaChat
# ============================================================

def create_gigachat(config):
    """Создать клиент GigaChat. Поддерживает 2 режима:
       1. credentials — прямой доступ (SberCloud)
       2. base_url + access_token — через прокси (закрытый контур)
    """
    from langchain_gigachat import GigaChat
    
    gc = config["gigachat"]
    model = gc.get("model", "GigaChat")
    
    # Режим 2: через прокси (base_url + access_token из env)
    if gc.get("base_url"):
        token_env = gc.get("access_token_env", "JPY_API_TOKEN")
        token = os.getenv(token_env, "")
        return GigaChat(
            base_url=gc["base_url"],
            access_token=token,
            model=model,
            timeout=gc.get("timeout", 60)
        )
    
    # Режим 1: через credentials
    if gc.get("credentials"):
        return GigaChat(
            credentials=gc["credentials"],
            model=model,
            verify_ssl_certs=gc.get("verify_ssl", False),
            timeout=gc.get("timeout", 60)
        )
    
    print("❌ Ошибка: Заполните gigachat.credentials или gigachat.base_url в config.yml")
    sys.exit(1)


# ============================================================
# Чтение структуры БД
# ============================================================

def get_db_connection(config):
    """Подключиться к PostgreSQL / GreenPlum"""
    if psycopg2 is None:
        print("❌ psycopg2 не установлен. Установите: pip install psycopg2-binary")
        print("   Или используйте --source duckdb / --source cube")
        sys.exit(1)
    db = config["database"]
    return psycopg2.connect(
        host=db["host"],
        port=db["port"],
        dbname=db["name"],
        user=db["user"],
        password=db["password"]
    )


def get_schema(config):
    """Получить имя схемы из конфига"""
    return config.get("database", {}).get("schema", "public")


def get_tables(conn, schema="public"):
    """Получить список таблиц"""
    cur = conn.cursor()
    cur.execute("""
        SELECT table_name 
        FROM information_schema.tables 
        WHERE table_schema = %s 
          AND table_type = 'BASE TABLE'
        ORDER BY table_name
    """, (schema,))
    tables = [row[0] for row in cur.fetchall()]
    cur.close()
    return tables


def get_columns(conn, table_name, schema="public"):
    """Получить колонки таблицы с типами"""
    cur = conn.cursor()
    cur.execute("""
        SELECT column_name, data_type, is_nullable, column_default,
               character_maximum_length
        FROM information_schema.columns
        WHERE table_schema = %s AND table_name = %s
        ORDER BY ordinal_position
    """, (schema, table_name))
    columns = []
    for row in cur.fetchall():
        columns.append({
            "name": row[0],
            "data_type": row[1],
            "nullable": row[2] == "YES",
            "default": row[3],
            "max_length": row[4]
        })
    cur.close()
    return columns


def get_foreign_keys(conn, table_name, schema="public"):
    """Получить внешние ключи таблицы"""
    cur = conn.cursor()
    cur.execute("""
        SELECT
            kcu.column_name,
            ccu.table_name AS foreign_table,
            ccu.column_name AS foreign_column
        FROM information_schema.table_constraints AS tc
        JOIN information_schema.key_column_usage AS kcu
            ON tc.constraint_name = kcu.constraint_name
        JOIN information_schema.constraint_column_usage AS ccu
            ON ccu.constraint_name = tc.constraint_name
        WHERE tc.constraint_type = 'FOREIGN KEY'
          AND tc.table_name = %s
          AND tc.table_schema = %s
    """, (table_name, schema))
    fks = []
    for row in cur.fetchall():
        fks.append({
            "column": row[0],
            "foreign_table": row[1],
            "foreign_column": row[2]
        })
    cur.close()
    return fks


def get_primary_key(conn, table_name, schema="public"):
    """Получить primary key таблицы"""
    cur = conn.cursor()
    cur.execute("""
        SELECT kcu.column_name
        FROM information_schema.table_constraints tc
        JOIN information_schema.key_column_usage kcu
            ON tc.constraint_name = kcu.constraint_name
        WHERE tc.constraint_type = 'PRIMARY KEY'
          AND tc.table_name = %s
          AND tc.table_schema = %s
    """, (table_name, schema))
    result = cur.fetchone()
    cur.close()
    return result[0] if result else "id"


def get_sample_data(conn, table_name, schema="public", limit=5):
    """Получить примеры данных из таблицы"""
    cur = conn.cursor()
    try:
        cur.execute(f'SELECT * FROM {schema}."{table_name}" LIMIT %s', (limit,))
        columns = [desc[0] for desc in cur.description]
        rows = cur.fetchall()
        cur.close()
        return columns, rows
    except Exception:
        cur.close()
        return [], []


def get_row_count(conn, table_name, schema="public"):
    """Получить количество строк"""
    cur = conn.cursor()
    cur.execute(f'SELECT COUNT(*) FROM {schema}."{table_name}"')
    count = cur.fetchone()[0]
    cur.close()
    return count


# ============================================================
# Чтение структуры из DuckDB
# ============================================================

class DuckDBSource:
    """Источник данных — локальный DuckDB-файл.
    Реализует тот же интерфейс что и psycopg2-функции выше.
    """

    def __init__(self, config):
        import duckdb
        db_path = config["database"].get("path", "./data.duckdb")
        self.schema = config["database"].get("schema", "main")
        self.conn = duckdb.connect(db_path, read_only=True)
        print(f"✅ DuckDB: {db_path} (schema={self.schema})")

    def get_tables(self):
        rows = self.conn.execute(
            "SELECT table_name FROM information_schema.tables "
            "WHERE table_schema = ? AND table_type = 'BASE TABLE' "
            "ORDER BY table_name",
            [self.schema]
        ).fetchall()
        return [r[0] for r in rows]

    def get_columns(self, table_name):
        rows = self.conn.execute(
            "SELECT column_name, data_type, is_nullable, column_default, "
            "character_maximum_length "
            "FROM information_schema.columns "
            "WHERE table_schema = ? AND table_name = ? "
            "ORDER BY ordinal_position",
            [self.schema, table_name]
        ).fetchall()
        return [{
            "name": r[0],
            "data_type": r[1],
            "nullable": r[2] == "YES",
            "default": r[3],
            "max_length": r[4]
        } for r in rows]

    def get_foreign_keys(self, table_name):
        # DuckDB поддерживает FK, но не всегда заполняет information_schema
        try:
            rows = self.conn.execute(
                "SELECT kcu.column_name, ccu.table_name, ccu.column_name "
                "FROM information_schema.table_constraints tc "
                "JOIN information_schema.key_column_usage kcu "
                "  ON tc.constraint_name = kcu.constraint_name "
                "JOIN information_schema.constraint_column_usage ccu "
                "  ON ccu.constraint_name = tc.constraint_name "
                "WHERE tc.constraint_type = 'FOREIGN KEY' "
                "  AND tc.table_name = ? AND tc.table_schema = ?",
                [table_name, self.schema]
            ).fetchall()
            return [{"column": r[0], "foreign_table": r[1], "foreign_column": r[2]}
                    for r in rows]
        except Exception:
            return []

    def get_primary_key(self, table_name):
        try:
            rows = self.conn.execute(
                "SELECT kcu.column_name "
                "FROM information_schema.table_constraints tc "
                "JOIN information_schema.key_column_usage kcu "
                "  ON tc.constraint_name = kcu.constraint_name "
                "WHERE tc.constraint_type = 'PRIMARY KEY' "
                "  AND tc.table_name = ? AND tc.table_schema = ?",
                [table_name, self.schema]
            ).fetchall()
            return rows[0][0] if rows else "id"
        except Exception:
            return "id"

    def get_sample_data(self, table_name, limit=5):
        try:
            result = self.conn.execute(
                f'SELECT * FROM {self.schema}."{table_name}" LIMIT ?', [limit]
            )
            columns = [desc[0] for desc in result.description]
            rows = result.fetchall()
            return columns, rows
        except Exception:
            return [], []

    def get_row_count(self, table_name):
        result = self.conn.execute(
            f'SELECT COUNT(*) FROM {self.schema}."{table_name}"'
        )
        return result.fetchone()[0]

    def close(self):
        self.conn.close()


# ============================================================
# Чтение метаданных из работающего Cube API
# ============================================================

class CubeAPISource:
    """Источник данных — метаданные из Cube REST API /meta.
    Не требует подключения к БД вообще.
    Cube уже работает → читаем его структуру.
    """

    def __init__(self, config):
        import httpx
        self.cube_url = config["cube"]["api_url"]
        self.schema = config["database"].get("schema", "public")
        headers = {}
        token = config["cube"].get("api_token", "")
        if token:
            headers["Authorization"] = f"Bearer {token}"

        resp = httpx.get(f"{self.cube_url}/meta", headers=headers, timeout=15.0)
        resp.raise_for_status()
        self.meta = resp.json()
        self.cubes = {c["name"]: c for c in self.meta.get("cubes", [])}
        print(f"✅ Cube API: {len(self.cubes)} кубов загружено из {self.cube_url}")

    def get_tables(self):
        return sorted(self.cubes.keys())

    def get_columns(self, table_name):
        cube = self.cubes.get(table_name, {})
        columns = []
        for dim in cube.get("dimensions", []):
            short_name = dim["name"].split(".")[-1]
            columns.append({
                "name": short_name,
                "data_type": dim.get("type", "string"),
                "nullable": True,
                "default": None,
                "max_length": None
            })
        return columns

    def get_foreign_keys(self, table_name):
        # Cube API не отдаёт FK — joins обнаружим по именам колонок
        return []

    def get_primary_key(self, table_name):
        cube = self.cubes.get(table_name, {})
        for dim in cube.get("dimensions", []):
            if dim.get("primaryKey"):
                return dim["name"].split(".")[-1]
        return "id"

    def get_sample_data(self, table_name, limit=5):
        # Не можем получить sample data через Cube API
        return [], []

    def get_row_count(self, table_name):
        # Пробуем count из Cube
        import httpx
        cube = self.cubes.get(table_name, {})
        count_measure = None
        for m in cube.get("measures", []):
            if m.get("type") == "count":
                count_measure = m["name"]
                break
        if not count_measure:
            return 0
        try:
            headers = {"Content-Type": "application/json"}
            resp = httpx.post(
                f"{self.cube_url}/load",
                json={"query": {"measures": [count_measure], "limit": 1}},
                headers=headers,
                timeout=15.0
            )
            data = resp.json().get("data", [])
            if data:
                return list(data[0].values())[0]
        except Exception:
            pass
        return 0

    def close(self):
        pass


# ============================================================
# Фабрика: выбор источника данных
# ============================================================

def create_data_source(config, override_source=None):
    """Создать источник данных на основе config.yml или --source аргумента.
    Возвращает объект с методами: get_tables, get_columns, get_foreign_keys,
    get_primary_key, get_sample_data, get_row_count, close.
    """
    driver = override_source or config.get("database", {}).get("driver", "postgresql")
    driver = driver.lower().strip()

    if driver == "duckdb":
        return DuckDBSource(config), driver
    elif driver == "cube":
        return CubeAPISource(config), driver
    elif driver in ("postgresql", "postgres"):
        conn = get_db_connection(config)
        schema = get_schema(config)
        return _PsycopgSource(conn, schema), driver
    elif driver == "greenplum":
        from db_sources import GreenplumSource
        return GreenplumSource(config), driver
    elif driver == "hive":
        from db_sources import HiveSource
        return HiveSource(config), driver
    else:
        print(f"❌ Неизвестный driver: {driver}")
        print("   Допустимые: postgresql, greenplum, hive, duckdb, cube")
        sys.exit(1)


class _PsycopgSource:
    """Обёртка над psycopg2 для единого интерфейса."""
    def __init__(self, conn, schema):
        self.conn = conn
        self.schema = schema

    def get_tables(self):
        return get_tables(self.conn, self.schema)

    def get_columns(self, table_name):
        return get_columns(self.conn, table_name, self.schema)

    def get_foreign_keys(self, table_name):
        return get_foreign_keys(self.conn, table_name, self.schema)

    def get_primary_key(self, table_name):
        return get_primary_key(self.conn, table_name, self.schema)

    def get_sample_data(self, table_name, limit=5):
        return get_sample_data(self.conn, table_name, self.schema, limit)

    def get_row_count(self, table_name):
        return get_row_count(self.conn, table_name, self.schema)

    def close(self):
        self.conn.close()


# ============================================================
# Knowledge Base — загрузка внешних подсказок для таблиц
# ============================================================

_KNOWLEDGE_BASE = {}  # Загружается из внешнего YAML-файла


def load_knowledge_base(kb_path: str) -> dict:
    """Загрузить Knowledge Base из YAML-файла.
    Возвращает dict: pattern_name → {title, description, column_hints, suggested_measures}.
    """
    global _KNOWLEDGE_BASE
    try:
        with open(kb_path, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f) or {}
        _KNOWLEDGE_BASE = data
        print(f"✅ Knowledge Base загружена: {len(data)} паттернов из {kb_path}")
        return data
    except FileNotFoundError:
        print(f"⚠️  KB файл не найден: {kb_path}")
        return {}
    except Exception as e:
        print(f"⚠️  Ошибка загрузки KB: {e}")
        return {}


def load_etl_plan(plan_path: str) -> dict:
    """Загрузить ETL execution plan файл (xlsx/csv).
    Возвращает dict: source_table_name → {columns from plan}.
    """
    info = {}
    plan_path_lower = plan_path.lower()

    try:
        if plan_path_lower.endswith(".xlsx") or plan_path_lower.endswith(".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(plan_path, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = dict(zip(headers, row))
                src_table = str(rec.get("source_table", "") or "").strip()
                if src_table:
                    info[src_table] = {
                        "source_schema": str(rec.get("source_schema", "") or ""),
                        "source_cluster": str(rec.get("source_cluster", "") or ""),
                        "target_table": str(rec.get("table_step2", "") or ""),
                        "process_description": str(rec.get("process_description", "") or "")[:500],
                        "last_updated": str(rec.get("last_updated_time", "") or ""),
                    }
        elif plan_path_lower.endswith(".csv"):
            import csv
            with open(plan_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for rec in reader:
                    src_table = rec.get("source_table", "").strip()
                    if src_table:
                        info[src_table] = {
                            "source_schema": rec.get("source_schema", ""),
                            "source_cluster": rec.get("source_cluster", ""),
                            "target_table": rec.get("table_step2", ""),
                            "process_description": rec.get("process_description", "")[:500],
                            "last_updated": rec.get("last_updated_time", ""),
                        }
        else:
            print(f"⚠️  Неподдерживаемый формат ETL plan: {plan_path}. Используйте .xlsx или .csv")
            return {}

        print(f"✅ ETL plan загружен: {len(info)} source-таблиц из {plan_path}")
        for t in info:
            print(f"   - {t}")
        return info

    except Exception as e:
        print(f"⚠️  Ошибка загрузки ETL plan: {e}")
        return {}


def _singularize(word: str) -> set:
    """Вернуть множество возможных единственных форм для английского слова."""
    forms = {word}
    if word.endswith("ies") and len(word) > 4:
        forms.add(word[:-3] + "y")          # priorities → priority
    if word.endswith("ses") or word.endswith("xes"):
        forms.add(word[:-2])                # statuses → status
    if word.endswith("es") and len(word) > 3:
        forms.add(word[:-2])                # statuses → status
    if word.endswith("s") and not word.endswith("ss"):
        forms.add(word[:-1])                # issues → issue
    return forms


def match_kb_hints(table_name: str, etl_plan: dict = None) -> dict:
    """Найти подсказки из Knowledge Base для таблицы.
    Сопоставление учитывает подчёркивания (issue_links ↔ issuelink),
    множественное число (priorities ↔ priority), префиксы (jiraissue ↔ issues).
    """
    tl = table_name.lower()
    tl_no_sep = tl.replace("_", "").replace("-", "")
    tl_singulars = _singularize(tl_no_sep)

    for pattern, hints in _KNOWLEDGE_BASE.items():
        pat_no_sep = pattern.replace("_", "")
        pat_singulars = _singularize(pat_no_sep)

        # 1. Точное совпадение (с нормализацией разделителей + числа)
        if pat_no_sep == tl_no_sep:
            return hints
        if tl_singulars & pat_singulars:
            return hints

        # 2. Префикс "jira" в KB: jiraissue → issue ↔ issues
        if pat_no_sep.startswith("jira"):
            core = pat_no_sep[4:]
            core_singulars = _singularize(core)
            if tl_singulars & core_singulars:
                return hints

        # 3. Префикс "project" в KB: projectversion → version ↔ versions
        if pat_no_sep.startswith("project"):
            core = pat_no_sep[7:]
            core_singulars = _singularize(core)
            if tl_singulars & core_singulars:
                return hints

        # 4. Суффиксный матч: dm_jira_components → component
        for pf in pat_singulars:
            if len(pf) >= 5 and any(tf.endswith(pf) for tf in tl_singulars):
                return hints

        # 5. Подстрока >= 6 символов (избегает ложных матчей)
        for pf in pat_singulars:
            if len(pf) >= 6 and pf in tl_no_sep:
                return hints

    if etl_plan:
        for src_table, plan_info in etl_plan.items():
            src_no_sep = src_table.lower().replace("_", "")
            src_singulars = _singularize(src_no_sep)
            if tl_singulars & src_singulars:
                return {
                    "title": f"Таблица из ETL ({src_table})",
                    "description": f"Источник: {plan_info.get('source_schema', '')}.{src_table}. "
                                   f"Целевая: {plan_info.get('target_table', '')}.",
                }

    return {}


def enrich_descriptions_with_kb(descriptions: dict, table_name: str,
                                columns: list, etl_plan: dict = None) -> dict:
    """Дополнить GigaChat-описания подсказками из Knowledge Base."""
    hints = match_kb_hints(table_name, etl_plan)
    if not hints:
        return descriptions

    col_descs = descriptions.get("columns", {})
    col_hints = hints.get("column_hints", {})

    for col_name, hint_text in col_hints.items():
        if col_name in col_descs:
            existing = col_descs[col_name]
            if not existing.get("description") or len(existing["description"]) < 10:
                existing["description"] = hint_text
        else:
            for c in columns:
                if c["name"].lower() == col_name or col_name in c["name"].lower():
                    col_descs[c["name"]] = col_descs.get(c["name"], {})
                    if not col_descs[c["name"]].get("description"):
                        col_descs[c["name"]]["description"] = hint_text
                    break

    descriptions["columns"] = col_descs
    return descriptions


def get_kb_suggested_measures(table_name: str) -> list:
    """Получить дополнительные меры из Knowledge Base."""
    hints = match_kb_hints(table_name)
    return hints.get("suggested_measures", [])


# ============================================================
# Обнаружение связей между таблицами
# ============================================================

def detect_implicit_relationships(table_name, columns, all_tables, explicit_fks):
    """
    Найти неявные связи по соглашению об именах (*_id → таблица).
    Обрабатывает:
      - Прямое совпадение: project_id → projects
      - Множественное число: priority_id → priorities, status_id → statuses
      - Префиксы доменных таблиц: status_id → issue_statuses, type_id → issue_types
      - Семантические маппинги: assignee_id → users, reporter_id → users, author_id → users
    """
    explicit_cols = {fk["column"] for fk in explicit_fks}
    implicit = []

    # Семантические маппинги: колонка → целевая таблица
    SEMANTIC_MAP = {
        "assignee": "users",
        "reporter": "users",
        "author": "users",
        "creator": "users",
        "owner": "users",
        "updated_by": "users",
        "created_by": "users",
        "lead": "users",
        "manager": "users",
        "parent": None,  # self-join обрабатывается отдельно
    }

    for col in columns:
        col_name = col["name"]
        if not col_name.endswith("_id") or col_name in explicit_cols:
            continue
        if col_name == "id":
            continue

        base = col_name[:-3]  # "project_id" → "project"

        # 1. Семантический маппинг
        if base in SEMANTIC_MAP:
            target = SEMANTIC_MAP[base]
            if target and target in all_tables:
                implicit.append({
                    "column": col_name,
                    "foreign_table": target,
                    "foreign_column": "id",
                    "source": "implicit"
                })
                continue
            elif target is None and table_name in all_tables:
                # self-join (parent_id → та же таблица)
                implicit.append({
                    "column": col_name,
                    "foreign_table": table_name,
                    "foreign_column": "id",
                    "source": "implicit"
                })
                continue

        # 2. Прямые кандидаты по имени
        candidates = [
            base + "s",       # project → projects
            base + "es",      # status → statuses
            base,             # sprint → sprint
        ]
        if base.endswith("y"):
            candidates.append(base[:-1] + "ies")  # priority → priorities, category → categories
        if base.endswith("s"):
            candidates.append(base)

        matched_table = None
        for candidate in candidates:
            if candidate in all_tables and candidate != table_name:
                matched_table = candidate
                break

        # 3. Если прямое не нашло — пробуем с доменными префиксами
        if not matched_table:
            # Получаем «домен» из имени текущей таблицы (issue_comments → issue)
            # и пробуем prefix_base (issue_status, issue_priority, etc.)
            prefixes_to_try = set()
            # Из имени таблицы: issues → issue, issue_comments → issue
            parts = table_name.split("_")
            if parts:
                singular = parts[0].rstrip("s")
                prefixes_to_try.add(singular)       # "issue"
                prefixes_to_try.add(parts[0])       # "issues"

            # Также общие доменные префиксы
            prefixes_to_try.update(["issue", "project", "workflow", "notification", "permission"])

            for prefix in prefixes_to_try:
                prefix_candidates = [
                    f"{prefix}_{base}s",       # issue_statuses
                    f"{prefix}_{base}es",      # issue_statuses
                    f"{prefix}_{base}",        # issue_type
                ]
                if base.endswith("y"):
                    prefix_candidates.append(f"{prefix}_{base[:-1]}ies")  # issue_priorities
                
                for candidate in prefix_candidates:
                    if candidate in all_tables and candidate != table_name:
                        matched_table = candidate
                        break
                if matched_table:
                    break

        # 4. Последняя попытка — поиск таблиц содержащих base в имени
        if not matched_table:
            for t in all_tables:
                if t != table_name and base in t and t.endswith("s"):
                    matched_table = t
                    break

        if matched_table:
            implicit.append({
                "column": col_name,
                "foreign_table": matched_table,
                "foreign_column": "id",
                "source": "implicit"
            })

    return implicit


def build_all_relationships(table_name, columns, all_tables, explicit_fks):
    """
    Объединить явные FK и неявные связи.
    При множественных ссылках на одну таблицу — генерировать алиасы.
    Возвращает список join-записей:
      [{"column": "assignee_id", "foreign_table": "users", "alias": "users_assignee",
        "foreign_column": "id", "relationship": "many_to_one"}]
    """
    # Собираем все связи
    all_rels = []
    for fk in explicit_fks:
        all_rels.append({**fk, "source": "explicit"})

    implicit = detect_implicit_relationships(table_name, columns, all_tables, explicit_fks)
    all_rels.extend(implicit)

    if not all_rels:
        return []

    # Считаем сколько раз каждая таблица фигурирует как цель
    target_count = {}
    for rel in all_rels:
        t = rel["foreign_table"]
        target_count[t] = target_count.get(t, 0) + 1

    # Генерируем записи с алиасами
    joins = []
    for rel in all_rels:
        target = rel["foreign_table"]
        col = rel["column"]
        base = col[:-3] if col.endswith("_id") else col  # assignee_id → assignee

        # Если таблица используется >1 раз — алиас обязателен
        if target_count[target] > 1:
            alias = f"{target}_{base}"  # users_assignee, users_reporter
        else:
            alias = target

        joins.append({
            "column": col,
            "foreign_table": target,
            "alias": alias,
            "foreign_column": rel.get("foreign_column", "id"),
            "relationship": "many_to_one",
            "source": rel.get("source", "explicit")
        })

    return joins


def _fix_missing_commas(text):
    """Вставить пропущенные запятые в JSON от GigaChat.
    Обрабатывает как многострочный, так и однострочный JSON.
    """
    import re
    # --- Многострочный: "value"\n  "key" → "value",\n  "key" ---
    text = re.sub(r'(")\s*\n(\s*")', r'\1,\n\2', text)
    text = re.sub(r'(})\s*\n(\s*")', r'\1,\n\2', text)
    text = re.sub(r'(\])\s*\n(\s*")', r'\1,\n\2', text)
    text = re.sub(r'(true|false|null|\d)\s*\n(\s*")', r'\1,\n\2', text)
    text = re.sub(r'(})\s*\n(\s*\{)', r'\1,\n\2', text)

    # --- Однострочный: "value" "key" → "value", "key" ---
    # "..." "..."  (два строковых значения подряд без запятой)
    text = re.sub(r'(") (")', r'\1, \2', text)
    # } "key"  →  }, "key"
    text = re.sub(r'(}) (")', r'\1, \2', text)
    # ] "key"  →  ], "key"
    text = re.sub(r'(\]) (")', r'\1, \2', text)
    # true/false/null/number  "key"
    text = re.sub(r'(true|false|null)(\s+)(")', r'\1,\2\3', text)
    text = re.sub(r'(\d)(\s+)(")', r'\1,\2\3', text)
    # } {  →  }, {  (массив объектов)
    text = re.sub(r'(})\s*(\{)', r'\1, \2', text)

    # --- Trailing commas ---
    text = re.sub(r',\s*}', '}', text)
    text = re.sub(r',\s*]', ']', text)
    return text


def _balance_brackets(text):
    """Добавить недостающие закрывающие скобки в JSON.
    GigaChat часто забывает одну или несколько } в конце ответа.
    """
    in_string = False
    escape = False
    opens = []
    match = {'{': '}', '[': ']'}
    for ch in text:
        if escape:
            escape = False
            continue
        if ch == '\\' and in_string:
            escape = True
            continue
        if ch == '"' and not escape:
            in_string = not in_string
            continue
        if in_string:
            continue
        if ch in ('{', '['):
            opens.append(ch)
        elif ch in ('}', ']'):
            if opens:
                opens.pop()
    closing = ''.join(match[o] for o in reversed(opens))
    if closing:
        text = text.rstrip()
        if text.endswith(','):
            text = text[:-1]
        text += closing
    return text


def _parse_json_safe(text):
    """Робастный парсинг JSON из ответа LLM.
    Обрабатывает: markdown-обёртки, типографские кавычки,
    пропущенные запятые, несбалансированные скобки.
    """
    import json as _json
    import re

    text = text.strip()

    # Убираем markdown
    if "```" in text:
        parts = text.split("```")
        for part in parts:
            part = part.strip()
            if part.startswith("json"):
                part = part[4:].strip()
            if part.startswith("{"):
                text = part
                break

    # Типографские кавычки
    for old, new in [('\u201c', '"'), ('\u201d', '"'), ('\u00ab', '"'), ('\u00bb', '"'),
                     ('\u2018', "'"), ('\u2019', "'")]:
        text = text.replace(old, new)

    # Извлекаем JSON-блок
    match = re.search(r'\{[\s\S]*\}', text)
    if match:
        text = match.group()

    # Попытка 1: как есть
    try:
        return _json.loads(text)
    except _json.JSONDecodeError:
        pass

    # Попытка 2: чиним пропущенные запятые
    fixed = _fix_missing_commas(text)
    try:
        return _json.loads(fixed)
    except _json.JSONDecodeError:
        pass

    # Попытка 3: балансировка скобок (GigaChat часто забывает закрывающие })
    balanced = _balance_brackets(fixed)
    try:
        return _json.loads(balanced)
    except _json.JSONDecodeError:
        pass

    # Попытка 4: агрессивная чистка — убираем невалидные символы
    cleaned = re.sub(r'[\x00-\x1f]+', ' ', balanced)
    for old, new in [('\u2014', '-'), ('\u2013', '-'), ('\u2026', '...')]:
        cleaned = cleaned.replace(old, new)
    try:
        return _json.loads(cleaned)
    except _json.JSONDecodeError as e:
        raise e


def suggest_joins_via_llm(llm, table_name, columns, detected_joins, all_tables):
    """
    Попросить GigaChat дать осмысленные описания для джойнов
    и предложить дополнительные связи, которые не были обнаружены автоматически.
    """
    if not detected_joins and not columns:
        return {}

    # Формируем текст уже найденных связей
    join_lines = []
    for j in detected_joins:
        join_lines.append(f"{j['column']} → {j['foreign_table']} (алиас: {j['alias']})")
    detected_text = "Связи:\n" + "\n".join(f"  - {l}" for l in join_lines)

    # Колонки с _id которые не попали в detected
    detected_cols = {j["column"] for j in detected_joins}
    unmatched_id_cols = [
        c["name"] for c in columns
        if c["name"].endswith("_id") and c["name"] != "id" and c["name"] not in detected_cols
    ]

    unmatched_text = ""
    if unmatched_id_cols:
        unmatched_text = f"\nНеразрешённые колонки: {', '.join(unmatched_id_cols)}"

    prompt = f"""Для каждой связи таблицы {table_name} дай title (1-2 слова) и description (1 предложение) на русском.

{detected_text}{unmatched_text}

Ответ строго в формате JSON:
{{"joins": [{{"column": "col_id", "title": "Название", "description": "Описание"}}], "extra_joins": []}}"""

    try:
        response = _llm_invoke_with_retry(llm, prompt)
        return _parse_json_safe(response.content)
    except Exception as e:
        print(f"  ⚠️ GigaChat не смог описать связи {table_name}: {e}")
        return {}


# ============================================================
# Генерация описаний через GigaChat
# ============================================================

def _llm_invoke_with_retry(llm, prompt, max_retries=3):
    """Вызов LLM с retry при rate-limit (429) и таймаутах."""
    import time as _time
    for attempt in range(max_retries):
        try:
            return llm.invoke(prompt)
        except Exception as e:
            err_str = str(e)
            if "429" in err_str or "Too Many Requests" in err_str or "timeout" in err_str.lower():
                wait = 5 * (attempt + 1)
                print(f"  ⏳ Rate limit / timeout, жду {wait}с (попытка {attempt+1}/{max_retries})...")
                _time.sleep(wait)
            else:
                raise
    raise RuntimeError(f"LLM не ответил после {max_retries} попыток")


def _analyze_sample_data(sample_columns, sample_rows, columns):
    """Статистический анализ sample-данных для обогащения промпта.
    Возвращает dict: col_name → {unique_values, null_count, value_type_hint}.
    """
    if not sample_rows or not sample_columns:
        return {}

    analysis = {}
    col_types = {c["name"]: c["data_type"] for c in columns}

    for idx, col_name in enumerate(sample_columns):
        values = [row[idx] for row in sample_rows]
        non_null = [v for v in values if v is not None]
        null_count = len(values) - len(non_null)

        info = {"null_count": null_count, "total": len(values)}

        if not non_null:
            info["hint"] = "все значения NULL"
            analysis[col_name] = info
            continue

        unique = set()
        for v in non_null:
            s = str(v).strip()
            if len(s) <= 100:
                unique.add(s)

        if len(unique) <= 10 and col_types.get(col_name, "") in (
            "character varying", "text", "varchar", "USER-DEFINED"
        ):
            info["unique_values"] = sorted(unique)
            info["hint"] = f"перечисление: {', '.join(sorted(unique))}"
        elif len(unique) <= 5:
            info["unique_values"] = sorted(unique)
            info["hint"] = f"примеры: {', '.join(sorted(unique)[:5])}"
        else:
            samples = [str(v)[:60] for v in non_null[:3]]
            info["hint"] = f"примеры: {', '.join(samples)}"

        if null_count > 0:
            pct = int(null_count / len(values) * 100)
            info["hint"] += f" ({pct}% NULL)"

        analysis[col_name] = info

    return analysis


def generate_descriptions(llm, table_name, columns, fks, sample_columns,
                          sample_rows, row_count, etl_context=None):
    """
    GigaChat: описания таблицы и колонок на основе структуры, примеров данных и ETL-контекста.
    """
    # Анализ sample data
    data_analysis = _analyze_sample_data(sample_columns, sample_rows, columns)

    # Формируем колонки с анализом
    col_lines = []
    for c in columns:
        line = f"  - {c['name']} ({c['data_type']})"
        hint = data_analysis.get(c["name"], {}).get("hint")
        if hint:
            line += f"  // {hint}"
        col_lines.append(line)
    columns_text = "\n".join(col_lines)

    # Примеры строк (компактно)
    sample_text = ""
    if sample_rows:
        sample_text = "\nПримеры строк:\n"
        for row in sample_rows[:3]:
            parts = []
            for col, val in zip(sample_columns, row):
                val_str = str(val)[:50] if val is not None else "NULL"
                parts.append(f"{col}={val_str}")
            sample_text += "  " + ", ".join(parts[:10]) + "\n"

    # FK-контекст
    fk_text = ""
    if fks:
        fk_lines = [f"  - {f['column']} → {f['foreign_table']}.{f['foreign_column']}" for f in fks]
        fk_text = "\nВнешние ключи:\n" + "\n".join(fk_lines)

    # ETL-контекст
    etl_text = ""
    if etl_context:
        parts = []
        if etl_context.get("process_description"):
            parts.append(f"Процесс: {etl_context['process_description']}")
        if etl_context.get("source_schema"):
            parts.append(f"Источник: {etl_context['source_schema']}")
        if etl_context.get("target_table"):
            parts.append(f"Целевая таблица ETL: {etl_context['target_table']}")
        if parts:
            etl_text = "\nETL-контекст: " + "; ".join(parts)

    prompt = f"""Опиши таблицу {table_name} ({row_count} строк) на русском.

Колонки (после // — реальные значения из данных):
{columns_text}
{sample_text}{fk_text}{etl_text}
ВАЖНО: Проанализируй реальные значения данных (примеры строк и значения после //).
Для колонок-перечислений (status, type, priority и пр.) обязательно перечисли допустимые значения в description.
Для числовых колонок укажи единицу измерения если можно определить из данных.

Ответ строго JSON:
{{"table_title": "Русское название (2-3 слова)", "table_description": "Описание (1-2 предложения)", "columns": {{"col_name": {{"title": "Название", "description": "Что хранит (с примерами значений)"}}}}}}"""

    # Попытка 1
    try:
        response = _llm_invoke_with_retry(llm, prompt)
        return _parse_json_safe(response.content)
    except Exception:
        pass

    # Попытка 2: сокращённый промпт с анализом
    col_hints = []
    for c in columns:
        hint = data_analysis.get(c["name"], {}).get("hint", "")
        col_hints.append(f"{c['name']}({c['data_type']}){': '+hint if hint else ''}")
    short_cols = "; ".join(col_hints[:20])

    prompt2 = f"""Таблица {table_name} ({row_count} строк). Колонки и данные: {short_cols}.
{etl_text}
Дай table_title (2 слова на русском) и table_description (1 предложение).
Для каждой колонки дай title (1-2 слова) и description (кратко, включая примеры значений).
Ответ: JSON {{"table_title":"...", "table_description":"...", "columns":{{"col":{{"title":"...", "description":"..."}}}}}}"""

    try:
        response = _llm_invoke_with_retry(llm, prompt2)
        return _parse_json_safe(response.content)
    except Exception as e:
        print(f"  ⚠️ GigaChat не смог описать {table_name}: {e}")

    # Fallback — описания на основе анализа данных (без LLM)
    result = {
        "table_title": table_name.replace("_", " ").title(),
        "table_description": f"Таблица {table_name}",
        "columns": {}
    }
    for c in columns:
        col_name = c["name"]
        hint = data_analysis.get(col_name, {}).get("hint", "")
        desc_text = hint if hint else f"{col_name} ({c['data_type']})"
        result["columns"][col_name] = {
            "title": col_name.replace("_", " ").replace("id", "").strip().title() or col_name,
            "description": desc_text
        }
    return result


# ============================================================
# Маппинг типов PostgreSQL → Cube.js
# ============================================================

def pg_type_to_cube(pg_type, column_name):
    """Сопоставить тип PostgreSQL с типом Cube.js"""
    pg_type = pg_type.lower()
    
    # Время
    if any(t in pg_type for t in ["timestamp", "date", "time"]):
        return "time"
    
    # Числа
    if any(t in pg_type for t in ["integer", "int", "bigint", "smallint", "serial",
                                    "numeric", "decimal", "real", "double", "float"]):
        return "number"
    
    # Булевы
    if "boolean" in pg_type:
        return "boolean"
    
    # Строки (по умолчанию)
    return "string"


# ============================================================
# Генерация Cube YAML
# ============================================================

def generate_cube_yaml(table_name, columns, enriched_joins, pk, descriptions,
                       schema="public", etl_context=None):
    """
    Сгенерировать YAML-модель Cube для одной таблицы.
    enriched_joins — результат build_all_relationships + описания от LLM.
    etl_context — данные из ETL execution plan для этой таблицы.
    """
    
    desc = descriptions
    cube_name = table_name
    
    table_desc = desc.get("table_description", "")
    if etl_context:
        etl_parts = []
        if etl_context.get("process_description"):
            etl_parts.append(etl_context["process_description"])
        if etl_context.get("source_schema"):
            etl_parts.append(f"Источник: {etl_context['source_schema']}")
        if etl_context.get("target_table"):
            etl_parts.append(f"ETL целевая: {etl_context['target_table']}")
        if etl_parts:
            etl_suffix = " | " + "; ".join(etl_parts)
            table_desc = table_desc.rstrip(".") + etl_suffix

    cube = {
        "name": cube_name,
        "sql_table": f"{schema}.{table_name}",
        "title": desc.get("table_title", table_name),
        "description": table_desc,
    }
    
    # --- Joins (обогащённые: FK + implicit + LLM) ---
    joins = []
    join_columns = set()  # колонки, задействованные в join
    
    for j in enriched_joins:
        alias = j["alias"]
        col = j["column"]
        foreign_col = j.get("foreign_column", "id")
        join_columns.add(col)
        
        join_entry = {
            "name": alias,
            "sql": "{CUBE}." + col + " = {" + alias + "}." + foreign_col,
            "relationship": j.get("relationship", "many_to_one"),
        }
        # Добавляем описание если есть
        if j.get("title"):
            join_entry["title"] = j["title"]
        if j.get("description"):
            join_entry["description"] = j["description"]
        
        joins.append(join_entry)
    
    if joins:
        cube["joins"] = joins
    
    # --- Dimensions ---
    dimensions = []
    col_descs = desc.get("columns", {})
    
    for c in columns:
        col_name = c["name"]
        cube_type = pg_type_to_cube(c["data_type"], col_name)
        
        # Пропускаем FK-колонки (они уходят через join)
        if col_name in join_columns and col_name != pk:
            continue
        
        dim = {
            "name": col_name,
            "sql": col_name,
            "type": cube_type,
        }
        
        if col_name == pk:
            dim["primary_key"] = True
        
        col_desc = col_descs.get(col_name, {})
        if col_desc.get("title"):
            dim["title"] = col_desc["title"]
        if col_desc.get("description"):
            dim["description"] = col_desc["description"]
        
        dimensions.append(dim)
    
    cube["dimensions"] = dimensions
    
    # --- Measures ---
    measures = [
        {
            "name": "count",
            "type": "count",
            "title": "Количество",
            "description": f"Общее количество записей в таблице {desc.get('table_title', table_name)}"
        }
    ]
    
    # Добавляем sum/avg для числовых колонок (не ID и не FK)
    for c in columns:
        col_name = c["name"]
        if col_name.endswith("_id") or col_name == pk:
            continue
        if pg_type_to_cube(c["data_type"], col_name) == "number":
            col_title = col_descs.get(col_name, {}).get("title", col_name)
            measures.append({
                "name": f"total_{col_name}",
                "sql": col_name,
                "type": "sum",
                "title": f"Сумма {col_title}",
                "description": f"Сумма значений поля {col_name}"
            })
            measures.append({
                "name": f"avg_{col_name}",
                "sql": col_name,
                "type": "avg",
                "title": f"Среднее {col_title}",
                "description": f"Среднее значение поля {col_name}"
            })
    
    # Дополнительные меры из Knowledge Base
    jira_measures = get_kb_suggested_measures(table_name)
    existing_names = {m["name"] for m in measures}
    for jm in jira_measures:
        if jm["name"] not in existing_names:
            measures.append(jm)

    cube["measures"] = measures
    
    return {"cubes": [cube]}


# ============================================================
# Генерация semantic-конфигов (glossary, examples)
# ============================================================

def generate_glossary(all_tables_info):
    """Сгенерировать базовый glossary.yml"""
    glossary = {}
    
    for info in all_tables_info:
        table = info["table_name"]
        desc = info["descriptions"]
        cube_name = table
        
        # Термин для таблицы
        title = desc.get("table_title", table)
        glossary[table] = {
            "aliases": [title.lower(), table, table.replace("_", " ")],
            "semantic_type": "entity",
            "fields": [f"{cube_name}.id"],
            "filter_operator": "equals",
            "description": desc.get("table_description", "")
        }
        
        # Термин count для таблицы
        glossary[f"{table}_count"] = {
            "aliases": [
                f"количество {title.lower()}",
                f"сколько {title.lower()}",
            ],
            "semantic_type": "metric",
            "measures": [f"{cube_name}.count"],
            "description": f"Количество записей {title}"
        }
    
    return glossary


def generate_examples(all_tables_info):
    """Сгенерировать базовый examples.yml"""
    examples = []
    
    for info in all_tables_info:
        table = info["table_name"]
        desc = info["descriptions"]
        cube_name = table
        title = desc.get("table_title", table)
        
        # Пример: "сколько <сущностей>"
        examples.append({
            "question": f"сколько {title.lower()}",
            "intent": "analytics",
            "query": {
                "measures": [f"{cube_name}.count"],
                "limit": 100
            },
            "tags": ["count", table]
        })
        
        # Пример: "список <сущностей>"
        dims = []
        for c in info["columns"][:5]:
            if c["name"] != "id" and not c["name"].endswith("_id"):
                dims.append(f"{cube_name}.{c['name']}")
        
        if dims:
            examples.append({
                "question": f"список {title.lower()}",
                "intent": "analytics",
                "query": {
                    "measures": [f"{cube_name}.count"],
                    "dimensions": dims[:4],
                    "limit": 100
                },
                "tags": ["list", table]
            })
    
    return examples


# ============================================================
# Парсинг Spark Execution Plan из ETL
# ============================================================

def _parse_spark_plan(process_description: str) -> dict:
    """Извлечь структурированную информацию из Spark execution plan.
    Возвращает: {source_tables, columns, joins, filters, target_table, target_columns}.
    """
    import re
    info = {
        "source_tables": [],
        "columns": [],
        "joins": [],
        "filters": [],
        "target_table": None,
        "target_columns": [],
    }
    if not process_description:
        return info

    text = process_description

    # Целевая таблица и колонки из INSERT
    m = re.search(r'InsertIntoHiveTable\s+`([^`]+)`\.`([^`]+)`.*?\[([^\]]*)\]', text)
    if m:
        info["target_table"] = f"{m.group(1)}.{m.group(2)}"
        cols_str = m.group(3)
        info["target_columns"] = [c.strip().split("=")[0].strip()
                                  for c in cols_str.split(",") if c.strip()]

    # CreateDataSourceTableAsSelectCommand
    m2 = re.search(r'CreateDataSourceTableAsSelectCommand\s+`([^`]+)`\.`([^`]+)`.*?\[([^\]]*)\]', text)
    if m2:
        info["target_table"] = f"{m2.group(1)}.{m2.group(2)}"
        info["target_columns"] = [c.strip() for c in m2.group(3).split(",") if c.strip()]

    # Scan hive → исходные таблицы
    for m in re.finditer(r'Scan hive\s+([^\s\[]+)\s*\[([^\]]*)\]', text):
        full_table = m.group(1)
        scan_cols = [c.strip().split("#")[0].strip() for c in m.group(2).split(",") if c.strip()]
        info["source_tables"].append({"table": full_table, "columns": scan_cols})
        info["columns"].extend(scan_cols)

    # Join-паттерны
    for m in re.finditer(r'(BroadcastHashJoin|SortMergeJoin)\s*\[([^\]]*)\],\s*\[([^\]]*)\]', text):
        left_col = m.group(2).strip().split("#")[0].strip()
        right_col = m.group(3).strip().split("#")[0].strip()
        info["joins"].append({"left": left_col, "right": right_col, "type": m.group(1)})

    # Фильтры
    for m in re.finditer(r'Filter\s*\((.+?)\)\s*$', text, re.MULTILINE):
        filt = m.group(1).strip()
        if len(filt) < 200:
            info["filters"].append(filt)

    info["columns"] = sorted(set(info["columns"]))
    return info


def _format_etl_summary(etl_entry: dict, parsed_plan: dict) -> str:
    """Форматировать ETL-информацию в читаемый текст для описания модели."""
    parts = []

    if parsed_plan.get("target_table"):
        parts.append(f"ETL целевая: {parsed_plan['target_table']}")
    if parsed_plan.get("target_columns"):
        parts.append(f"Колонки: {', '.join(parsed_plan['target_columns'][:15])}")
    if parsed_plan.get("source_tables"):
        src_names = list({s["table"].split(".")[-1] for s in parsed_plan["source_tables"]})
        parts.append(f"Источники: {', '.join(src_names[:10])}")
    if parsed_plan.get("joins"):
        join_descs = [f"{j['left']}↔{j['right']}" for j in parsed_plan["joins"][:5]]
        parts.append(f"Связи: {', '.join(join_descs)}")

    if etl_entry.get("source_schema"):
        parts.append(f"Схема: {etl_entry['source_schema']}")
    if etl_entry.get("last_updated"):
        parts.append(f"Обновлено: {etl_entry['last_updated']}")

    return " | ".join(parts)


# ============================================================
# Обогащение существующих моделей через ETL plan
# ============================================================

def enrich_models_with_etl(model_dir: str, etl_plan: dict, llm=None,
                           data_source=None, kb_path: str = None) -> dict:
    """Обогатить уже сгенерированные Cube YAML-модели данными из ETL plan.

    Параметры:
      model_dir   — папка с .yml файлами моделей Cube
      etl_plan    — dict из load_etl_plan()
      llm         — GigaChat (опционально, для переописания колонок)
      data_source — источник данных (опционально, для sample data)
      kb_path     — путь к KB (опционально)

    Возвращает: {updated: [...], skipped: [...], errors: [...]}
    """
    model_path = Path(model_dir)
    if not model_path.exists():
        print(f"❌ Папка моделей не найдена: {model_dir}")
        return {"updated": [], "skipped": [], "errors": []}

    if kb_path and Path(kb_path).exists():
        load_knowledge_base(kb_path)

    # Парсим все ETL-записи заранее
    etl_parsed = {}
    for src_table, entry in etl_plan.items():
        parsed = _parse_spark_plan(entry.get("process_description", ""))
        etl_parsed[src_table] = {"entry": entry, "parsed": parsed}

    yml_files = sorted(model_path.glob("*.yml"))
    print(f"\n📂 Моделей в {model_dir}: {len(yml_files)}")
    print(f"📋 Записей в ETL plan: {len(etl_plan)}")
    print()

    results = {"updated": [], "skipped": [], "errors": []}

    for yml_file in yml_files:
        cube_name = yml_file.stem
        try:
            with open(yml_file, 'r', encoding='utf-8') as f:
                model = yaml.safe_load(f)
        except Exception as e:
            results["errors"].append(f"{cube_name}: ошибка чтения — {e}")
            continue

        if not model or "cubes" not in model or not model["cubes"]:
            results["skipped"].append(f"{cube_name}: нет блока cubes")
            continue

        cube = model["cubes"][0]

        # Сопоставляем с ETL plan
        matched_key = None
        matched_data = None
        cube_norm = cube_name.lower().replace("_", "")
        cube_singulars = _singularize(cube_norm)

        for src_table, data in etl_parsed.items():
            src_norm = src_table.lower().replace("_", "")
            src_singulars = _singularize(src_norm)

            if cube_singulars & src_singulars:
                matched_key = src_table
                matched_data = data
                break

            # Проверяем целевую таблицу
            target = data["parsed"].get("target_table", "")
            if target:
                target_short = target.split(".")[-1].lower().replace("_", "")
                target_singulars = _singularize(target_short)
                if cube_singulars & target_singulars:
                    matched_key = src_table
                    matched_data = data
                    break

            # Проверяем исходные таблицы в плане
            for st in data["parsed"].get("source_tables", []):
                st_short = st["table"].split(".")[-1].lower().replace("_", "")
                st_singulars = _singularize(st_short)
                if cube_singulars & st_singulars:
                    matched_key = src_table
                    matched_data = data
                    break
            if matched_key:
                break

        if not matched_data:
            results["skipped"].append(cube_name)
            continue

        print(f"🔗 {cube_name} ← ETL: {matched_key}")

        entry = matched_data["entry"]
        parsed = matched_data["parsed"]

        etl_summary = _format_etl_summary(entry, parsed)
        changes = []

        # 1. Обогащаем description куба
        old_desc = cube.get("description", "")
        if "ETL" not in old_desc and etl_summary:
            cube["description"] = (old_desc.rstrip(". ") + " | " + etl_summary) if old_desc else etl_summary
            changes.append("description")

        # 2. Обогащаем колонки из parsed plan
        if parsed.get("target_columns"):
            dim_names = {d["name"] for d in cube.get("dimensions", [])}
            etl_cols = set(parsed["target_columns"])
            new_cols_in_etl = etl_cols - dim_names
            if new_cols_in_etl:
                changes.append(f"ETL-колонки не в модели: {', '.join(sorted(new_cols_in_etl))}")

        # 3. Если есть LLM + data_source — переописываем колонки с ETL-контекстом
        if llm and data_source:
            try:
                columns = data_source.get_columns(cube_name)
                fks = data_source.get_foreign_keys(cube_name)
                row_count = data_source.get_row_count(cube_name)
                scols, srows = data_source.get_sample_data(cube_name, 10)

                new_desc = generate_descriptions(
                    llm, cube_name, columns, fks, scols, srows, row_count,
                    etl_context=entry
                )

                # Обновляем title/description если GigaChat дал лучше
                if new_desc.get("table_title") and len(new_desc["table_title"]) > len(cube.get("title", "")):
                    cube["title"] = new_desc["table_title"]
                    changes.append("title (GigaChat+ETL)")

                if new_desc.get("table_description") and len(new_desc["table_description"]) > 20:
                    cube["description"] = new_desc["table_description"]
                    if etl_summary and "ETL" not in cube["description"]:
                        cube["description"] += " | " + etl_summary
                    changes.append("description (GigaChat+ETL)")

                # Обновляем описания dimensions
                col_descs = new_desc.get("columns", {})
                for dim in cube.get("dimensions", []):
                    dname = dim["name"]
                    new_col = col_descs.get(dname, {})
                    if new_col.get("title") and dim.get("title", "") in (dname, dname.replace("_", " ").title(), ""):
                        dim["title"] = new_col["title"]
                    old_dim_desc = dim.get("description", "")
                    if new_col.get("description") and (
                        not old_dim_desc or old_dim_desc.endswith(")") or len(old_dim_desc) < 15
                    ):
                        dim["description"] = new_col["description"]

                changes.append("dimensions (GigaChat+ETL)")

            except Exception as e:
                changes.append(f"⚠️ GigaChat: {e}")

        # 4. Обогащаем описания из Knowledge Base
        if _KNOWLEDGE_BASE:
            kb_hints = match_kb_hints(cube_name)
            if kb_hints:
                col_hints = kb_hints.get("column_hints", {})
                for dim in cube.get("dimensions", []):
                    dname = dim["name"]
                    if dname in col_hints:
                        old_dim_desc = dim.get("description", "")
                        if not old_dim_desc or old_dim_desc.endswith(")") or len(old_dim_desc) < 15:
                            dim["description"] = col_hints[dname]

                # Добавляем suggested measures
                suggested = kb_hints.get("suggested_measures", [])
                existing_measure_names = {m["name"] for m in cube.get("measures", [])}
                for sm in suggested:
                    if sm["name"] not in existing_measure_names:
                        cube.setdefault("measures", []).append(sm)
                        changes.append(f"measure: {sm['name']}")

        if changes:
            model["cubes"][0] = cube
            with open(yml_file, 'w', encoding='utf-8') as f:
                yaml.dump(model, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
            results["updated"].append(f"{cube_name}: {', '.join(changes)}")
            print(f"   ✅ {', '.join(changes)}")
        else:
            results["skipped"].append(cube_name)
            print(f"   ⏭️  Нет изменений")

    print(f"\n{'='*60}")
    print(f"  Обновлено: {len(results['updated'])}")
    print(f"  Пропущено: {len(results['skipped'])}")
    print(f"  Ошибок:    {len(results['errors'])}")
    print(f"{'='*60}\n")

    return results


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Загрузчик данных в Cube")
    parser.add_argument("--source", choices=["postgresql", "greenplum", "hive", "duckdb", "cube"],
                        help="Переопределить database.driver из config.yml")
    parser.add_argument("--kb", metavar="FILE",
                        help="Путь к Knowledge Base YAML (переопределяет config.yml)")
    parser.add_argument("--etl-plan", metavar="FILE",
                        help="Путь к ETL execution plan (xlsx/csv) для обогащения моделей")
    parser.add_argument("--enrich-etl", action="store_true",
                        help="Обогатить существующие модели через ETL plan (без перегенерации)")
    parser.add_argument("--enrich-with-llm", action="store_true",
                        help="При --enrich-etl переописывать колонки через GigaChat + sample data")
    parser.add_argument("--model-dir", metavar="DIR",
                        help="Папка с моделями (для --enrich-etl, по умолчанию из config.yml)")
    args = parser.parse_args()

    # 1. Загрузить конфиг
    config = load_config()

    # ── Режим: обогащение существующих моделей через ETL plan ──
    if args.enrich_etl:
        print("=" * 60)
        print("  ОБОГАЩЕНИЕ МОДЕЛЕЙ ЧЕРЕЗ ETL PLAN")
        print("=" * 60)
        print()

        plan_file = args.etl_plan or config.get("etl_plan_path")
        if not plan_file:
            print("❌ Укажите ETL plan: --etl-plan <file.xlsx> или etl_plan_path в config.yml")
            sys.exit(1)
        if not Path(plan_file).exists():
            print(f"❌ ETL plan файл не найден: {plan_file}")
            sys.exit(1)

        etl_plan = load_etl_plan(plan_file)
        if not etl_plan:
            print("❌ ETL plan пустой или не удалось загрузить")
            sys.exit(1)

        model_dir = args.model_dir or config.get("cube", {}).get("model_path", "./cube_models")
        kb_path = args.kb or config.get("knowledge_base_path")

        llm = None
        data_source = None
        if args.enrich_with_llm:
            print("🔄 Подключение к GigaChat...")
            llm = create_gigachat(config)
            print("✅ GigaChat готов")
            data_source, _ = create_data_source(config, args.source)
            print(f"✅ Источник данных подключён")

        results = enrich_models_with_etl(
            model_dir=model_dir,
            etl_plan=etl_plan,
            llm=llm,
            data_source=data_source,
            kb_path=kb_path
        )

        if data_source:
            data_source.close()

        print("Следующие шаги:")
        print("  1. Проверьте обновлённые модели в папке моделей")
        print("  2. Перезапустите Cube: npx cubejs-server")
        print("  3. Пересоберите FAISS: python 02_build_faiss.py")
        return

    # ── Режим: полная генерация моделей ──
    print("=" * 60)
    print("  ЗАГРУЗЧИК ДАННЫХ В CUBE")
    print("  Источник → Описания GigaChat → YAML-модели Cube")
    print("=" * 60)
    print()
    
    print("✅ Конфигурация загружена")
    
    # 2. Подключиться к источнику данных
    source, driver_name = create_data_source(config, args.source)
    schema = config.get("database", {}).get("schema", "public")
    print(f"   Источник: {driver_name}, Схема: {schema}")
    
    # 3. Загрузить Knowledge Base (если указана)
    kb_path = args.kb or config.get("knowledge_base_path")
    if kb_path and Path(kb_path).exists():
        load_knowledge_base(kb_path)
    elif kb_path:
        print(f"⚠️  KB файл не найден: {kb_path}")
    else:
        print("ℹ️  Knowledge Base не указана (--kb или knowledge_base_path в config.yml)")

    # 4. Загрузить ETL plan (если указан)
    etl_plan = {}
    plan_file = args.etl_plan or config.get("etl_plan_path")
    if plan_file and Path(plan_file).exists():
        etl_plan = load_etl_plan(plan_file)
    elif plan_file:
        print(f"⚠️  ETL plan файл не найден: {plan_file}")

    # 5. Подключить GigaChat
    print("🔄 Подключение к GigaChat...")
    llm = create_gigachat(config)
    print("✅ GigaChat готов")
    
    # 6. Получить список таблиц
    tables = source.get_tables()
    print(f"\n📋 Найдено таблиц: {len(tables)}")
    for t in tables:
        print(f"   - {t}")
    print()
    
    # 5. Обработать каждую таблицу
    model_path = Path(config["cube"]["model_path"])
    model_path.mkdir(parents=True, exist_ok=True)
    
    all_tables_set = set(tables)
    all_tables_info = []
    
    for i, table in enumerate(tables, 1):
        print(f"[{i}/{len(tables)}] Обработка таблицы: {table}")
        
        # Читаем структуру через унифицированный интерфейс
        columns = source.get_columns(table)
        fks = source.get_foreign_keys(table)
        pk = source.get_primary_key(table)
        row_count = source.get_row_count(table)
        sample_cols, sample_rows = source.get_sample_data(table, 5)
        
        print(f"   Колонок: {len(columns)}, FK: {len(fks)}, Строк: {row_count}")
        
        # Обнаруживаем все связи (FK + implicit по именам)
        enriched_joins = build_all_relationships(table, columns, all_tables_set, fks)
        implicit_count = sum(1 for j in enriched_joins if j.get("source") == "implicit")
        if enriched_joins:
            print(f"   🔗 Связей: {len(enriched_joins)} (FK: {len(fks)}, по именам: {implicit_count})")
            for j in enriched_joins:
                src = "FK" if j["source"] == "explicit" else "→"
                print(f"      {src} {j['column']} → {j['foreign_table']} (as {j['alias']})")
        
        # Просим GigaChat описать связи (если они есть)
        if enriched_joins:
            print(f"   🤖 GigaChat: описание связей...")
            join_suggestions = suggest_joins_via_llm(llm, table, columns, enriched_joins, all_tables_set)
            
            # Обогащаем joins описаниями от LLM
            llm_joins_map = {}
            for lj in join_suggestions.get("joins", []):
                key = lj.get("column", "")
                llm_joins_map[key] = lj
            
            for j in enriched_joins:
                llm_info = llm_joins_map.get(j["column"], {})
                if llm_info.get("title"):
                    j["title"] = llm_info["title"]
                if llm_info.get("description"):
                    j["description"] = llm_info["description"]
                if llm_info.get("alias") and llm_info["alias"] != j["alias"]:
                    j["alias"] = llm_info["alias"]
            
            # Добавляем extra_joins от LLM
            for extra in join_suggestions.get("extra_joins", []):
                extra_table = extra.get("foreign_table", "")
                if extra_table in all_tables_set and extra_table != table:
                    col_names = {c["name"] for c in columns}
                    if extra.get("column") in col_names:
                        enriched_joins.append({
                            "column": extra["column"],
                            "foreign_table": extra_table,
                            "alias": extra.get("alias", extra_table),
                            "foreign_column": extra.get("foreign_column", "id"),
                            "relationship": "many_to_one",
                            "title": extra.get("title", ""),
                            "description": extra.get("description", ""),
                            "source": "llm"
                        })
                        print(f"      ✨ LLM предложил: {extra['column']} → {extra_table}")
        
        # ETL-контекст для текущей таблицы
        etl_context = None
        if etl_plan:
            for src_name, plan_info in etl_plan.items():
                src_norm = src_name.lower().replace("_", "")
                table_norm = table.lower().replace("_", "")
                if src_norm == table_norm or table_norm in src_norm or src_norm in table_norm:
                    etl_context = plan_info
                    print(f"   📋 ETL plan: сопоставлена с {src_name}")
                    break

        # Генерируем описания через GigaChat
        print(f"   🤖 GigaChat: описания таблицы и колонок...")
        descriptions = generate_descriptions(
            llm, table, columns, fks, sample_cols, sample_rows, row_count,
            etl_context=etl_context
        )

        # Обогащаем описания из Knowledge Base
        kb_hints = match_kb_hints(table, etl_plan)
        if kb_hints:
            print(f"   📚 KB: {kb_hints.get('title', 'match found')}")
            descriptions = enrich_descriptions_with_kb(descriptions, table, columns, etl_plan)
            if not descriptions.get("table_description") or len(descriptions["table_description"]) < 10:
                descriptions["table_description"] = kb_hints.get("description", descriptions.get("table_description", ""))
            if not descriptions.get("table_title") or descriptions["table_title"] == table:
                descriptions["table_title"] = kb_hints.get("title", descriptions.get("table_title", table))

        print(f"   ✅ Описания: {descriptions.get('table_title', '?')}")
        
        # Генерируем Cube YAML
        cube_schema = schema if driver_name != "duckdb" else "main"
        cube_yaml = generate_cube_yaml(table, columns, enriched_joins, pk, descriptions,
                                        cube_schema, etl_context)
        
        # Сохраняем
        yaml_path = model_path / f"{table}.yml"
        with open(yaml_path, 'w', encoding='utf-8') as f:
            yaml.dump(cube_yaml, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
        
        print(f"   💾 Сохранено: {yaml_path}")
        
        all_tables_info.append({
            "table_name": table,
            "columns": columns,
            "fks": fks,
            "enriched_joins": enriched_joins,
            "descriptions": descriptions
        })
        print()
    
    # 6. Генерируем semantic-конфиги
    config_path = Path("config")
    config_path.mkdir(exist_ok=True)
    
    print("📝 Генерация glossary.yml...")
    glossary = generate_glossary(all_tables_info)
    with open(config_path / "glossary.yml", 'w', encoding='utf-8') as f:
        yaml.dump(glossary, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
    
    print("📝 Генерация examples.yml...")
    examples = generate_examples(all_tables_info)
    with open(config_path / "examples.yml", 'w', encoding='utf-8') as f:
        yaml.dump(examples, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
    
    print("📝 Генерация semantic_layer.yml...")
    layer_config = {
        "cube": {
            "base_url": config["cube"]["api_url"],
            "enabled": True,
            "preferred_cubes": [t["table_name"] for t in all_tables_info]
        },
        "intents": {
            "analytics": {
                "description": "Аналитические запросы через Cube",
                "keywords": ["сколько", "количество", "покажи", "список",
                             "средний", "сумма", "топ", "всего", "по проектам"],
                "priority": 1
            }
        },
        "query_generation": {
            "default_limit": 100,
            "max_limit": 10000
        }
    }
    with open(config_path / "semantic_layer.yml", 'w', encoding='utf-8') as f:
        yaml.dump(layer_config, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
    
    source.close()
    
    # Сводка по связям
    total_joins = sum(len(info.get("enriched_joins", [])) for info in all_tables_info)
    fk_joins = sum(
        sum(1 for j in info.get("enriched_joins", []) if j.get("source") == "explicit")
        for info in all_tables_info
    )
    implicit_joins = sum(
        sum(1 for j in info.get("enriched_joins", []) if j.get("source") == "implicit")
        for info in all_tables_info
    )
    llm_joins = sum(
        sum(1 for j in info.get("enriched_joins", []) if j.get("source") == "llm")
        for info in all_tables_info
    )
    
    print()
    print("=" * 60)
    print("  ✅ ГОТОВО!")
    print("=" * 60)
    print(f"""
Источник данных: {driver_name}
Созданные файлы:
  📁 {model_path}/          — YAML-модели Cube ({len(tables)} файлов)
  📁 config/glossary.yml    — Бизнес-глоссарий
  📁 config/examples.yml    — Примеры запросов
  📁 config/semantic_layer.yml — Конфигурация семантического слоя

Обнаруженные связи (joins):
  🔗 Всего: {total_joins}
     - Из FK constraints: {fk_joins}
     - По именам колонок: {implicit_joins}
     - Предложено GigaChat: {llm_joins}

Следующие шаги:
  1. Скопируйте файлы из {model_path}/ в папку model/cubes/ вашего Cube-проекта
  2. ПРОВЕРЬТЕ СВЯЗИ: откройте YAML-файлы и убедитесь что joins корректны
  3. Перезапустите Cube: npx cubejs-server
  4. Проверьте: curl http://localhost:4000/cubejs-api/v1/meta
  5. Отредактируйте glossary.yml и examples.yml под ваши задачи
  6. Запустите 02_build_faiss.py для создания векторного индекса
  7. Откройте 03_agent.ipynb в JupyterLab и задавайте вопросы
""")


if __name__ == "__main__":
    main()
